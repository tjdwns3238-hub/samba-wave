"""AI 소싱기 서비스 - 근거데이터 기반 최적 상품 조합 생성.

무신사 인기상품, 네이버 데이터랩 쇼핑인사이트, 판매 엑셀 데이터를 분석하여
IP안전 브랜드 × 키워드 조합을 자동 생성하고 검색그룹으로 등록한다.

핵심: "카테고리"가 아닌 "키워드"로 조합을 만든다.
  - "아디다스 패션잡화" (X) → 카테고리 매핑 불가
  - "아디다스 운동화" (O) → 검색그룹에서 카테고리 매핑 가능
브랜드×키워드 조합은 실제 데이터에서 발견된 쌍만 사용한다.
  - 크록스×운동화 (O) — 실제 존재
  - 크록스×목도리 (X) — 논리적 불가
"""

from __future__ import annotations

import asyncio
import io
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Optional

import httpx

from backend.domain.samba.proxy.musinsa import MusinsaClient
from backend.utils import now_kst
from backend.utils.logger import logger


# ── 무신사 카테고리 매핑 ──

MUSINSA_CATEGORIES = {
    "상의": "001",
    "아우터": "002",
    "바지": "003",
    "원피스/스커트": "020",
    "가방": "004",
    "신발": "005",
    "시계/주얼리": "006",
    "패션소품": "007",
    "스포츠/레저": "017",
    "속옷/슬립웨어": "026",
}

# 네이버 데이터랩 쇼핑인사이트 카테고리
NAVER_DL_CATEGORIES = {
    "패션의류": "50000000",
    "패션잡화": "50000001",
    "화장품/미용": "50000002",
    "스포츠/레저": "50000007",
    "생활/건강": "50000008",
}

# ── 대카테고리 → 무신사/네이버 자동 매핑 ──
MAIN_CATEGORIES: dict[str, dict[str, list[str]]] = {
    "패션의류": {
        "musinsa": ["상의", "아우터", "바지", "원피스/스커트", "속옷/슬립웨어"],
        "naver": ["패션의류"],
    },
    "패션잡화": {
        "musinsa": ["가방", "신발", "시계/주얼리", "패션소품"],
        "naver": ["패션잡화"],
    },
    "스포츠/레저": {
        "musinsa": ["스포츠/레저"],
        "naver": ["스포츠/레저"],
    },
    "패션전체": {
        "musinsa": list(MUSINSA_CATEGORIES.keys()),
        "naver": ["패션의류", "패션잡화"],
    },
}

# ── 카테고리별 검색 키워드 매핑 ──
# 무신사 카테고리 코드로 검색하여 브랜드를 추출하고,
# 상품명에서 구체적 키워드를 함께 추출한다.
CATEGORY_KEYWORDS: dict[str, list[str]] = {
    "상의": [
        "티셔츠",
        "맨투맨",
        "후드",
        "니트",
        "셔츠",
        "블라우스",
        "가디건",
        "조끼",
        "반팔",
        "긴팔",
        "롱슬리브",
        "크루넥",
        "폴로",
        "스웨터",
        "후드티",
        "탱크탑",
        "크롭",
        "오버핏",
        "반팔티",
        "긴팔티",
    ],
    "아우터": [
        "자켓",
        "코트",
        "패딩",
        "점퍼",
        "바람막이",
        "야상",
        "블레이저",
        "후리스",
        "트랙탑",
        "블루종",
        "후드집업",
        "집업",
        "트레이닝자켓",
        "윈드브레이커",
        "가죽자켓",
        "데님자켓",
        "트러커",
        "헤링턴",
        "봄버",
    ],
    "바지": [
        "청바지",
        "슬랙스",
        "조거팬츠",
        "숏팬츠",
        "카고팬츠",
        "레깅스",
        "트레이닝팬츠",
        "와이드팬츠",
        "스웨트팬츠",
        "데님팬츠",
        "반바지",
        "숏츠",
        "치노",
        "밴딩팬츠",
        "원턱",
        "투턱",
    ],
    "원피스/스커트": [
        "원피스",
        "스커트",
        "미니스커트",
        "롱스커트",
        "미디스커트",
        "플리츠",
    ],
    "가방": [
        "백팩",
        "크로스백",
        "토트백",
        "숄더백",
        "에코백",
        "클러치",
        "웨이스트백",
        "메신저백",
        "보스턴백",
    ],
    "신발": [
        "운동화",
        "스니커즈",
        "슬리퍼",
        "샌들",
        "로퍼",
        "부츠",
        "구두",
        "런닝화",
        "발렛",
        "뮬",
        "슈즈",
        "트레킹화",
        "캔버스화",
        "데크슈즈",
        "첼시부츠",
        "워커",
    ],
    "시계/주얼리": ["시계", "반지", "목걸이", "팔찌", "귀걸이", "체인", "펜던트"],
    "패션소품": [
        "모자",
        "머플러",
        "장갑",
        "벨트",
        "양말",
        "넥타이",
        "스카프",
        "캡",
        "볼캡",
        "버킷햇",
        "비니",
        "삭스",
        "헤어밴드",
        "선글라스",
    ],
    "스포츠/레저": [
        "트레이닝복",
        "래쉬가드",
        "등산화",
        "요가복",
        "수영복",
        "트레이닝",
        "축구화",
        "농구화",
        "골프",
    ],
    "속옷/슬립웨어": ["속옷", "팬티", "브라", "파자마", "잠옷", "홈웨어"],
    "뷰티": [
        "틴트",
        "립스틱",
        "파운데이션",
        "향수",
        "쿠션",
        "아이섀도",
        "마스카라",
        "선크림",
    ],
}

# 전체 카테고리 키워드 평탄화 (모듈 로드 시 1회만 계산)
_ALL_CATEGORY_KEYWORDS: list[str] = list(
    set(kw for kws in CATEGORY_KEYWORDS.values() for kw in kws)
)


@dataclass
class BrandScore:
    """브랜드별 점수."""

    brand: str
    count: int = 0
    total_sales: float = 0
    avg_profit_rate: float = 0
    score: float = 0
    categories: list[str] = field(default_factory=list)
    keywords: list[str] = field(default_factory=list)  # 실제 발견된 구체적 키워드
    source: str = ""


@dataclass
class BrandKeywordPair:
    """브랜드×키워드 실제 발견 쌍. 조합 생성의 기본 단위."""

    brand: str
    keyword: str
    count: int = 0  # 해당 쌍의 출현 빈도
    score: float = 0  # 해당 쌍의 점수
    source: str = ""  # musinsa / naver / excel


@dataclass
class Combination:
    """소싱처×브랜드×키워드 조합."""

    source_site: str
    brand: str
    keyword: str  # 구체적 키워드 (운동화, 슬리퍼 등)
    category: str  # 무신사 대분류 (신발, 상의 등) — 참고용
    category_code: str
    estimated_count: int
    search_url: str
    is_safe: bool = True
    safety_reason: str = ""


# 영문 → 한글 브랜드 매핑 (영문 수집 시 한글로 통합)
BRAND_EN_TO_KR: dict[str, str] = {
    "nike": "나이키",
    "adidas": "아디다스",
    "new balance": "뉴발란스",
    "puma": "푸마",
    "crocs": "크록스",
    "skechers": "스케쳐스",
    "vans": "반스",
    "converse": "컨버스",
    "reebok": "리복",
    "the north face": "노스페이스",
    "patagonia": "파타고니아",
    "columbia": "컬럼비아",
    "asics": "아식스",
    "mizuno": "미즈노",
    "salomon": "살로몬",
    "hoka": "호카",
    "on running": "온러닝",
    "gucci": "구찌",
    "louis vuitton": "루이비통",
    "chanel": "샤넬",
    "prada": "프라다",
    "celine": "셀린느",
    "balenciaga": "발렌시아가",
    "dior": "디올",
    "hermes": "에르메스",
    "burberry": "버버리",
    "coach": "코치",
    "michael kors": "마이클코어스",
    "spao": "스파오",
    "jeep": "지프",
    "snowpeak": "스노우피크",
    "casio": "카시오",
    "umbro": "엄브로",
    "champion": "챔피온",
    "tommy hilfiger": "타미힐피거",
    "calvin klein": "캘빈클라인",
    "lacoste": "라코스테",
    "polo": "폴로",
    "diesel": "디젤",
    "levis": "리바이스",
    "zara": "자라",
    "h&m": "H&M",
    "fila": "필라",
    "stussy": "스투시",
    "carhartt": "칼하트",
    "dickies": "디키즈",
    "kolon sport": "코오롱스포츠",
    "dynafit": "다이나핏",
    "noiago": "노이아고",
    "rawrow": "로우로우",
    "charles & keith": "찰스앤키스",
    "jansport": "잔스포츠",
    "descente": "데상트",
    "le coq": "르꼬끄",
    "k2": "K2",
    "blackyak": "블랙야크",
    "eider": "아이더",
    "nepa": "네파",
    "beanpole": "빈폴",
    "hazzys": "해지스",
    "mlb": "MLB",
    "sisley": "시슬리",
    "hugo boss": "휴고보스",
    "armani": "아르마니",
    "marc jacobs": "마크제이콥스",
    "tory burch": "토리버치",
    "nationalgeographic": "내셔널지오그래픽",
}


def normalize_brand(brand: str) -> str:
    """영문 브랜드를 한글로 통합. 매핑 없으면 원본 반환."""
    return BRAND_EN_TO_KR.get(brand.lower().strip(), brand)


class AISourcingService:
    """AI 소싱기 - 근거데이터 분석 및 검색그룹 생성."""

    def __init__(self, musinsa_cookie: str = ""):
        self.musinsa = MusinsaClient(cookie=musinsa_cookie)

    # ══════════════════════════════════════════
    # 1. 무신사 인기상품에서 브랜드 + 키워드 추출
    # ══════════════════════════════════════════

    async def fetch_musinsa_popular(
        self,
        categories: list[str] | None = None,
        size: int = 200,
    ) -> tuple[list[BrandScore], list[BrandKeywordPair]]:
        """무신사 키워드별 인기상품에서 브랜드 + 구체적 키워드 추출.

        무신사 API는 keyword 필수 → 카테고리별 대표 키워드로 각각 검색.
        예: "신발" 카테고리 → "운동화", "슬리퍼", "샌들" 등으로 각각 검색.
        Returns: (브랜드 목록, 브랜드×키워드 쌍 목록)
        """
        if not categories:
            categories = list(MUSINSA_CATEGORIES.keys())

        brand_map: dict[str, BrandScore] = {}
        pair_map: dict[str, BrandKeywordPair] = {}  # key: "brand|keyword"

        for cat_name in categories:
            cat_keywords = CATEGORY_KEYWORDS.get(cat_name, [])
            if not cat_keywords:
                continue

            # 카테고리별 대표 키워드로 각각 검색 (category 파라미터 제거 — 일부 코드 400 에러)
            for search_kw in cat_keywords:
                try:
                    result = await self.musinsa.search_products(
                        keyword=search_kw,
                        sort="POPULAR",
                        size=min(size, 200),
                    )
                    items = result.get("data", [])

                    for item in items:
                        brand = item.get("brand", "").strip()
                        if not brand:
                            continue
                        price = item.get("salePrice") or item.get("originalPrice") or 0

                        # 브랜드 집계
                        if brand not in brand_map:
                            brand_map[brand] = BrandScore(
                                brand=brand, source="musinsa", categories=[]
                            )
                        bs = brand_map[brand]
                        bs.count += 1
                        bs.total_sales += price
                        if cat_name not in bs.categories:
                            bs.categories.append(cat_name)

                        # 검색 키워드 = 곧 브랜드×키워드 쌍
                        if search_kw not in bs.keywords:
                            bs.keywords.append(search_kw)
                        pair_key = f"{brand}|{search_kw}"
                        if pair_key not in pair_map:
                            pair_map[pair_key] = BrandKeywordPair(
                                brand=brand, keyword=search_kw, source="musinsa"
                            )
                        pair_map[pair_key].count += 1
                        pair_map[pair_key].score += price / 10000

                    logger.info(
                        f"[AI소싱] 무신사 '{search_kw}'({cat_name}) → {len(items)}개 수집"
                    )
                except Exception as e:
                    logger.warning(f"[AI소싱] 무신사 '{search_kw}' 수집 실패: {e}")

        # 점수 계산
        for bs in brand_map.values():
            avg_price = bs.total_sales / bs.count if bs.count > 0 else 0
            bs.score = bs.count * (avg_price / 10000)

        brand_list = sorted(brand_map.values(), key=lambda x: x.score, reverse=True)
        pair_list = sorted(pair_map.values(), key=lambda x: x.score, reverse=True)
        return brand_list, pair_list

    def extract_brands_from_ranking(
        self,
        ranking_items: list[dict],
    ) -> tuple[list[BrandScore], list[BrandKeywordPair]]:
        """확장앱이 수집한 랭킹 아카이브 DOM 데이터에서 브랜드+키워드 추출.

        각 item은 { goodsNo, texts: [텍스트들] } 형태.
        texts[0]이 보통 브랜드, texts[1]이 상품명.
        """
        all_cat_keywords = _ALL_CATEGORY_KEYWORDS

        brand_map: dict[str, BrandScore] = {}
        pair_map: dict[str, BrandKeywordPair] = {}

        for item in ranking_items:
            # 새 구조: { rank, brand, name, price, goodsNo }
            brand = item.get("brand", "").strip()
            product_name = item.get("name", "").strip()
            price = item.get("price", 0) or 0

            # 이전 구조 호환: { texts: [...] }
            if not brand and "texts" in item:
                texts = item.get("texts", [])
                for t in texts:
                    if isinstance(t, str) and not t.isdigit() and not brand:
                        brand = t.strip()
                    elif (
                        isinstance(t, str)
                        and not t.isdigit()
                        and brand
                        and not product_name
                    ):
                        product_name = t.strip()
                        break

            if not brand:
                continue

            # 브랜드 집계
            if brand not in brand_map:
                brand_map[brand] = BrandScore(brand=brand, source="ranking")
            bs = brand_map[brand]
            bs.count += 1
            bs.total_sales += price

            # 상품명에서 키워드 추출
            if product_name:
                found_kws = self._extract_keywords_from_name(
                    product_name, all_cat_keywords
                )
                for kw in found_kws:
                    if kw not in bs.keywords:
                        bs.keywords.append(kw)
                    pair_key = f"{brand}|{kw}"
                    if pair_key not in pair_map:
                        pair_map[pair_key] = BrandKeywordPair(
                            brand=brand, keyword=kw, source="ranking"
                        )
                    pair_map[pair_key].count += 1
                    pair_map[pair_key].score += price / 10000 if price else 1

        # 점수 계산
        for bs in brand_map.values():
            avg_price = bs.total_sales / bs.count if bs.count > 0 else 0
            bs.score = bs.count * (avg_price / 10000) if avg_price > 0 else bs.count

        brand_list = sorted(brand_map.values(), key=lambda x: x.score, reverse=True)
        pair_list = sorted(pair_map.values(), key=lambda x: x.score, reverse=True)
        return brand_list, pair_list

    # 영문→한국어 키워드 매핑 (상품명이 영문일 때 한국어 키워드로 변환)
    _EN_KW_MAP: dict[str, str] = {
        "track top": "트랙탑",
        "track jacket": "트랙탑",
        "track pants": "트레이닝팬츠",
        "sweat pants": "스웨트팬츠",
        "sweatpants": "스웨트팬츠",
        "hoodie": "후드",
        "hood": "후드",
        "zip up": "집업",
        "zip-up": "집업",
        "jacket": "자켓",
        "blazer": "블레이저",
        "coat": "코트",
        "t-shirt": "티셔츠",
        "tee": "티셔츠",
        "tshirt": "티셔츠",
        "shirts": "셔츠",
        "shirt": "셔츠",
        "polo": "폴로",
        "pants": "바지",
        "trousers": "슬랙스",
        "slacks": "슬랙스",
        "denim": "청바지",
        "jeans": "청바지",
        "shorts": "반바지",
        "short pants": "반바지",
        "sneakers": "스니커즈",
        "sneaker": "스니커즈",
        "running": "런닝화",
        "runner": "런닝화",
        "boots": "부츠",
        "boot": "부츠",
        "sandal": "샌들",
        "sandals": "샌들",
        "slipper": "슬리퍼",
        "slippers": "슬리퍼",
        "slide": "슬리퍼",
        "loafer": "로퍼",
        "loafers": "로퍼",
        "ballet": "발렛",
        "mule": "뮬",
        "backpack": "백팩",
        "bag": "가방",
        "tote": "토트백",
        "cap": "캡",
        "hat": "모자",
        "bucket": "버킷햇",
        "beanie": "비니",
        "socks": "양말",
        "belt": "벨트",
        "cardigan": "가디건",
        "knit": "니트",
        "sweater": "스웨터",
        "vest": "조끼",
        "windbreaker": "바람막이",
        "blouson": "블루종",
        "bomber": "봄버",
        "tint": "틴트",
        "lipstick": "립스틱",
    }

    @classmethod
    def _extract_keywords_from_name(
        cls, product_name: str, category_keywords: list[str]
    ) -> list[str]:
        """상품명에서 구체적 키워드 추출.

        예: "나이키 에어맥스 97 런닝화" → ["런닝화"]
        예: "Deep One Tuck Sweat Pants" → ["스웨트팬츠"]
        """
        name_lower = product_name.lower().replace(" ", "")
        found: list[str] = []
        # 한국어 키워드 직접 매칭
        for kw in category_keywords:
            if kw.lower().replace(" ", "") in name_lower:
                found.append(kw)
        # 영문 키워드 매핑
        if not found:
            name_en = product_name.lower()
            for en_kw, ko_kw in cls._EN_KW_MAP.items():
                if en_kw in name_en and ko_kw not in found:
                    found.append(ko_kw)
        return found

    # ══════════════════════════════════════════
    # 2. 네이버 데이터랩 쇼핑인사이트 인기검색어
    # ══════════════════════════════════════════

    async def fetch_naver_datalab(
        self,
        naver_categories: list[str] | None = None,
        month: int = 0,
    ) -> tuple[list[BrandScore], list[BrandKeywordPair]]:
        """네이버 데이터랩에서 인기검색어 추출 → 브랜드+키워드 쌍 파싱.

        "나이키운동화" → 브랜드=나이키, 키워드=운동화
        Returns: (브랜드 목록, 브랜드×키워드 쌍 목록)
        """
        if not naver_categories:
            naver_categories = ["패션의류", "패션잡화"]

        all_keywords: list[dict[str, Any]] = []

        for cat_name in naver_categories:
            cid = NAVER_DL_CATEGORIES.get(cat_name)
            if not cid:
                continue
            try:
                keywords = await self._scrape_naver_datalab_keywords(
                    cid, cat_name, month=month
                )
                all_keywords.extend(keywords)
                logger.info(
                    f"[AI소싱] 네이버 데이터랩 {cat_name} → {len(keywords)}개 키워드"
                )
            except Exception as e:
                logger.warning(f"[AI소싱] 네이버 데이터랩 {cat_name} 실패: {e}")

        # 모든 카테고리의 키워드 목록 통합
        all_cat_keywords = _ALL_CATEGORY_KEYWORDS

        brand_map: dict[str, BrandScore] = {}
        pair_map: dict[str, BrandKeywordPair] = {}
        known_brands = self._get_known_brand_list()

        for kw_info in all_keywords:
            search_term = kw_info.get("keyword", "")
            rank = kw_info.get("rank", 999)
            category = kw_info.get("category", "")

            # 검색어에서 브랜드 추출
            brand = self._extract_brand_from_keyword(search_term, known_brands)
            if not brand:
                continue

            # 검색어에서 상품 키워드 추출 (브랜드 부분 제거 후)
            product_keyword = self._extract_product_keyword(
                search_term, brand, all_cat_keywords
            )

            # 브랜드 집계
            if brand not in brand_map:
                brand_map[brand] = BrandScore(
                    brand=brand, source="naver", categories=[]
                )
            bs = brand_map[brand]
            bs.count += 1
            bs.score += max(501 - rank, 1)
            if category and category not in bs.categories:
                bs.categories.append(category)

            # 브랜드×키워드 쌍
            if product_keyword:
                if product_keyword not in bs.keywords:
                    bs.keywords.append(product_keyword)
                pair_key = f"{brand}|{product_keyword}"
                if pair_key not in pair_map:
                    pair_map[pair_key] = BrandKeywordPair(
                        brand=brand, keyword=product_keyword, source="naver"
                    )
                pair_map[pair_key].count += 1
                pair_map[pair_key].score += max(501 - rank, 1)

        brand_list = sorted(brand_map.values(), key=lambda x: x.score, reverse=True)
        pair_list = sorted(pair_map.values(), key=lambda x: x.score, reverse=True)
        return brand_list, pair_list

    @staticmethod
    def _extract_product_keyword(
        search_term: str, brand: str, known_keywords: list[str]
    ) -> str:
        """검색어에서 브랜드를 제거하고 상품 키워드를 추출.

        "나이키운동화" → "운동화"
        "아디다스 슬리퍼" → "슬리퍼"
        "뉴발란스993" → "" (키워드 없음)
        """
        # 브랜드 제거
        remaining = search_term.replace(brand, "").strip()
        remaining_lower = remaining.lower().replace(" ", "")

        # 알려진 키워드 매칭 (긴 것부터)
        sorted_kws = sorted(known_keywords, key=len, reverse=True)
        for kw in sorted_kws:
            if kw.lower().replace(" ", "") in remaining_lower:
                return kw

        return ""

    async def _scrape_naver_datalab_keywords(
        self, cid: str, category_name: str, month: int = 0
    ) -> list[dict[str, Any]]:
        """네이버 데이터랩 쇼핑인사이트 인기검색어 스크래핑."""
        now = now_kst()
        if month > 0:
            import calendar

            target_year = now.year - 1
            last_day = calendar.monthrange(target_year, month)[1]
            start_date = f"{target_year}-{month:02d}-01"
            end_date = f"{target_year}-{month:02d}-{last_day:02d}"
        else:
            end_date = now.strftime("%Y-%m-%d")
            month_ago = now.month - 1 if now.month > 1 else 12
            year_ago = now.year if now.month > 1 else now.year - 1
            start_date = f"{year_ago}-{month_ago:02d}-{now.day:02d}"

        url = "https://datalab.naver.com/shoppingInsight/getKeywordRank.naver"
        headers = {
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Referer": "https://datalab.naver.com/shoppingInsight/sCategory.naver",
            "Content-Type": "application/x-www-form-urlencoded",
        }
        form_data = {
            "cid": cid,
            "timeUnit": "date",
            "startDate": start_date,
            "endDate": end_date,
            "age": "",
            "gender": "",
            "device": "",
        }

        timeout = httpx.Timeout(15.0, connect=10.0)
        async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
            resp = await client.post(url, data=form_data, headers=headers)
            resp.raise_for_status()
            data = resp.json()

        keywords: list[dict[str, Any]] = []
        seen: set[str] = set()

        if isinstance(data, list):
            for day in data:
                for item in day.get("ranks", []):
                    kw = item.get("keyword", "")
                    if kw and kw not in seen:
                        seen.add(kw)
                        keywords.append(
                            {
                                "keyword": kw,
                                "rank": item.get("rank", 999),
                                "category": category_name,
                            }
                        )
        elif isinstance(data, dict):
            for item in data.get("ranks", []):
                kw = item.get("keyword", "")
                if kw and kw not in seen:
                    seen.add(kw)
                    keywords.append(
                        {
                            "keyword": kw,
                            "rank": item.get("rank", 999),
                            "category": category_name,
                        }
                    )

        return keywords

    def _get_known_brand_list(self) -> list[str]:
        """알려진 브랜드 목록."""
        return [
            # 글로벌
            "나이키",
            "Nike",
            "아디다스",
            "Adidas",
            "뉴발란스",
            "New Balance",
            "푸마",
            "Puma",
            "크록스",
            "Crocs",
            "스케쳐스",
            "Skechers",
            "반스",
            "Vans",
            "컨버스",
            "Converse",
            "리복",
            "Reebok",
            "노스페이스",
            "The North Face",
            "파타고니아",
            "Patagonia",
            "컬럼비아",
            "Columbia",
            "아식스",
            "Asics",
            "미즈노",
            "Mizuno",
            "살로몬",
            "Salomon",
            "호카",
            "Hoka",
            "온러닝",
            "On Running",
            # 명품 (IP위험)
            "구찌",
            "Gucci",
            "루이비통",
            "Louis Vuitton",
            "샤넬",
            "Chanel",
            "프라다",
            "Prada",
            "셀린느",
            "Celine",
            "발렌시아가",
            "Balenciaga",
            "디올",
            "Dior",
            "에르메스",
            "Hermes",
            "버버리",
            "Burberry",
            "코치",
            "Coach",
            "마이클코어스",
            "Michael Kors",
            # 국내
            "무신사 스탠다드",
            "무신사스탠다드",
            "꼼파뇨",
            "디스커버리",
            "내셔널지오그래픽",
            "MLB",
            "폴햄",
            "탑텐",
            "스파오",
            "SPAO",
            "지프",
            "JEEP",
            "르무르",
            "스노우피크",
            "SNOWPEAK",
            "노르디스크",
            "엘브로",
            "카시오",
            "Casio",
            "다니엘웰링턴",
            "앤더슨벨",
            "커버낫",
            "이미스",
            "마르디메크르디",
            "엄브로",
            "Umbro",
            "챔피온",
            "Champion",
            "타미힐피거",
            "Tommy Hilfiger",
            "캘빈클라인",
            "Calvin Klein",
            "라코스테",
            "Lacoste",
            "폴로",
            "Polo",
            "디젤",
            "Diesel",
            "리바이스",
            "Levis",
            "지오다노",
            "유니클로",
            "자라",
            "ZARA",
            "H&M",
            "필라",
            "FILA",
            "휠라",
            "꼼데가르송",
            "아크네",
            "메종키츠네",
            "스투시",
            "Stussy",
            "칼하트",
            "Carhartt",
            "디키즈",
            "Dickies",
            # 국내 추가
            "코오롱스포츠",
            "KOLON SPORT",
            "코오롱",
            "다이나핏",
            "DYNAFIT",
            "노이아고",
            "NOIAGO",
            "로우로우",
            "RAWROW",
            "에고이스트",
            "써스데이아일랜드",
            "올리비아로렌",
            "제시뉴욕",
            "찰스앤키스",
            "Charles & Keith",
            "잔스포츠",
            "JanSport",
            "데상트",
            "DESCENTE",
            "르꼬끄",
            "Le Coq",
            "K2",
            "케이투",
            "블랙야크",
            "BLACKYAK",
            "아이더",
            "EIDER",
            "네파",
            "NEPA",
            "빈폴",
            "BEANPOLE",
            "해지스",
            "HAZZYS",
            "시슬리",
            "SISLEY",
            "마인",
            "MINE",
            "랩",
            "LAP",
            "에잇세컨즈",
            "8seconds",
            "쿠론",
            "COURONNE",
            "MCM",
            "닥스",
            "DAKS",
            "휴고보스",
            "Hugo Boss",
            "아르마니",
            "Armani",
            "마크제이콥스",
            "Marc Jacobs",
            "토리버치",
            "Tory Burch",
        ]

    def _extract_brand_from_keyword(self, keyword: str, known_brands: list[str]) -> str:
        """검색어에서 브랜드명 추출."""
        keyword_lower = keyword.lower()
        sorted_brands = sorted(known_brands, key=len, reverse=True)
        for brand in sorted_brands:
            if brand.lower() in keyword_lower:
                return brand
        return ""

    # ══════════════════════════════════════════
    # 3. 판매 엑셀 파싱
    # ══════════════════════════════════════════

    def parse_sales_excel(
        self, file_content: bytes, month: int = 0
    ) -> list[BrandScore]:
        """판매 엑셀에서 브랜드별 매출/수익률 집계.

        month > 0이면 해당월 데이터만 필터링한다.
        시트 자동 탐색: 판매 관련 시트명 우선, 없으면 '상품명' 헤더가 있는 시트 선택.
        """
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)

        # 시트 자동 탐색: 판매 관련 시트명 우선
        sales_sheet_patterns = [
            "샵마인",
            "판매",
            "주문",
            "매출",
            "정산",
            "sales",
            "order",
        ]
        ws = None
        # 1) 판매 관련 시트명 검색
        for sheet_name in wb.sheetnames:
            name_lower = sheet_name.lower().strip()
            for pat in sales_sheet_patterns:
                if pat.lower() in name_lower:
                    ws = wb[sheet_name]
                    logger.info(
                        f"[AI소싱] 엑셀 시트 자동선택: '{sheet_name}' (패턴: {pat})"
                    )
                    break
            if ws:
                break
        # 2) 못 찾으면 모든 시트에서 '상품명' 헤더 탐색
        if ws is None:
            for sheet_name in wb.sheetnames:
                candidate = wb[sheet_name]
                for row_idx in range(1, min(candidate.max_row or 10, 10) + 1):
                    for col_idx in range(1, min(candidate.max_column or 30, 30) + 1):
                        val = str(
                            candidate.cell(row=row_idx, column=col_idx).value or ""
                        ).strip()
                        if val in ("상품명", "상품", "품명", "product"):
                            ws = candidate
                            logger.info(
                                f"[AI소싱] 엑셀 시트 자동선택: '{sheet_name}' (헤더 발견)"
                            )
                            break
                    if ws:
                        break
                if ws:
                    break
        # 3) 그래도 못 찾으면 활성 시트
        if ws is None:
            ws = wb.active
        if ws is None:
            return []

        logger.info(
            f"[AI소싱] 엑셀 시트: '{ws.title}', 총 {ws.max_row}행 × {ws.max_column}열"
        )

        # 헤더 행 찾기 (날짜 컬럼 포함)
        header_row = None
        col_map: dict[str, int] = {}
        # "번호", "코드" 등이 포함된 헤더는 "상품명"에서 제외
        exclude_if_contains = {"상품명": ["번호", "코드", "no", "id", "code"]}
        target_headers = {
            "상품명": ["상품명", "품명", "product"],
            "매출": ["정산예상금", "정산금액", "매출액", "매출", "판매금액", "금액"],
            "수익률": ["수익률", "이익률", "마진율", "margin"],
            "수익": ["수익금", "수익", "이익", "profit", "순이익"],
            "날짜": [
                "날짜",
                "주문일",
                "결제일",
                "발주일",
                "주문일자",
                "결제일자",
                "date",
                "order_date",
            ],
        }

        for row_idx in range(1, min(ws.max_row or 10, 10) + 1):
            row_values = []
            for col_idx in range(1, min(ws.max_column or 30, 30) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = str(cell.value or "").strip().lower()
                row_values.append((col_idx, val))

            for col_idx, val in row_values:
                for key, patterns in target_headers.items():
                    if key in col_map:
                        continue
                    # 제외 패턴 확인
                    excluded = False
                    for ex in exclude_if_contains.get(key, []):
                        if ex.lower() in val:
                            excluded = True
                            break
                    if excluded:
                        continue
                    for pat in patterns:
                        if pat.lower() in val:
                            col_map[key] = col_idx
                            break

            if "상품명" in col_map:
                header_row = row_idx
                break

        if not header_row or "상품명" not in col_map:
            logger.warning("[AI소싱] 엑셀에서 '상품명' 컬럼을 찾을 수 없음")
            return []

        logger.info(f"[AI소싱] 엑셀 헤더 행={header_row}, 컬럼매핑={col_map}")
        if month > 0 and "날짜" not in col_map:
            logger.warning("[AI소싱] 날짜 컬럼 없음 → 월 필터링 없이 전체 처리")

        known_brands = self._get_known_brand_list()
        brand_data: dict[str, dict[str, Any]] = defaultdict(
            lambda: {"count": 0, "total_sales": 0.0, "profit_rates": [], "keywords": []}
        )

        # 키워드 추출용
        all_cat_keywords = _ALL_CATEGORY_KEYWORDS

        filtered_count = 0
        total_count = 0

        for row_idx in range(header_row + 1, (ws.max_row or 0) + 1):
            product_name = str(
                ws.cell(row=row_idx, column=col_map["상품명"]).value or ""
            ).strip()
            if not product_name:
                continue
            total_count += 1

            # 날짜 필터링: month > 0이고 날짜 컬럼이 있으면 해당월만 처리
            if month > 0 and "날짜" in col_map:
                date_val = ws.cell(row=row_idx, column=col_map["날짜"]).value
                if not self._match_month(date_val, month):
                    continue
                filtered_count += 1

            brand = self._extract_brand_from_product_name(product_name, known_brands)
            if not brand:
                continue

            bd = brand_data[brand]
            bd["count"] += 1

            # 키워드 추출
            found_kws = self._extract_keywords_from_name(product_name, all_cat_keywords)
            for kw in found_kws:
                if kw not in bd["keywords"]:
                    bd["keywords"].append(kw)

            if "매출" in col_map:
                sales_val = ws.cell(row=row_idx, column=col_map["매출"]).value
                if sales_val and isinstance(sales_val, (int, float)):
                    bd["total_sales"] += float(sales_val)

            if "수익률" in col_map:
                rate_val = ws.cell(row=row_idx, column=col_map["수익률"]).value
                if rate_val:
                    try:
                        rate = float(str(rate_val).replace("%", "").strip())
                        if rate > 1:
                            rate = rate / 100
                        bd["profit_rates"].append(rate)
                    except (ValueError, TypeError):
                        pass

        if month > 0 and "날짜" in col_map:
            logger.info(
                f"[AI소싱] 엑셀 {month}월 필터: {total_count}행 중 {filtered_count}행 매칭"
            )

        result: list[BrandScore] = []
        for brand, bd in brand_data.items():
            avg_rate = (
                sum(bd["profit_rates"]) / len(bd["profit_rates"])
                if bd["profit_rates"]
                else 0
            )
            sales_weight = (
                min(bd["total_sales"] / 100000, 10) if bd["total_sales"] > 0 else 1
            )
            score = bd["count"] * (1 + avg_rate) * sales_weight

            result.append(
                BrandScore(
                    brand=brand,
                    count=bd["count"],
                    total_sales=bd["total_sales"],
                    avg_profit_rate=avg_rate,
                    score=score,
                    keywords=bd["keywords"],
                    source="excel",
                )
            )

        return sorted(result, key=lambda x: x.score, reverse=True)

    def _match_month(self, date_val: Any, target_month: int) -> bool:
        """셀 값에서 월을 추출하여 target_month와 일치하는지 확인."""
        if date_val is None:
            return False
        # datetime 객체인 경우
        if isinstance(date_val, datetime):
            return date_val.month == target_month
        # 문자열인 경우 (2025-03-15, 2025/03/15, 25.03.15 등)
        s = str(date_val).strip()
        if not s:
            return False
        # yyyy-mm-dd, yyyy/mm/dd 형태
        m = re.match(r"(\d{4})[/\-.](\d{1,2})[/\-.](\d{1,2})", s)
        if m:
            return int(m.group(2)) == target_month
        # yy-mm-dd 형태
        m = re.match(r"(\d{2})[/\-.](\d{1,2})[/\-.](\d{1,2})", s)
        if m:
            return int(m.group(2)) == target_month
        return False

    def _extract_brand_from_product_name(
        self, name: str, known_brands: list[str]
    ) -> str:
        """상품명에서 브랜드 추출."""
        # 공통 접두사 제거
        prefixes = ["매장정품", "정품", "해외직구", "당일발송", "무료배송"]
        cleaned = name
        for prefix in prefixes:
            if cleaned.startswith(prefix):
                cleaned = cleaned[len(prefix) :].strip()

        name_lower = cleaned.lower()
        sorted_brands = sorted(known_brands, key=len, reverse=True)
        for brand in sorted_brands:
            if brand.lower() in name_lower:
                return brand

        tokens = cleaned.split()
        if tokens:
            first = tokens[0].strip()
            if len(first) >= 2 and re.match(r"^[가-힣a-zA-Z]+$", first):
                return first
        return ""

    # ══════════════════════════════════════════
    # 4. IP안전 브랜드 필터링
    # ══════════════════════════════════════════

    async def check_brand_safety(
        self,
        brands: list[str],
        forbidden_words: list[str],
    ) -> dict[str, dict[str, Any]]:
        """금지어 DB 기반 브랜드 IP안전 검증."""
        result: dict[str, dict[str, Any]] = {}
        forbidden_lower = [w.lower() for w in forbidden_words]

        for brand in brands:
            brand_lower = brand.lower()
            is_safe = True
            reason = "안전"

            for fw in forbidden_lower:
                if fw in brand_lower or brand_lower in fw:
                    is_safe = False
                    reason = f"금지어 매칭: {fw}"
                    break

            result[brand] = {"is_safe": is_safe, "reason": reason}

        return result

    # ══════════════════════════════════════════
    # 5. 조합 생성 — 브랜드×키워드 쌍 기반
    # ══════════════════════════════════════════

    async def generate_combinations(
        self,
        safe_brands: list[BrandScore],
        brand_keyword_pairs: list[BrandKeywordPair],
        target_count: int = 0,
        existing_combos: set[str] | None = None,
    ) -> list[Combination]:
        """실제 발견된 브랜드×키워드 쌍으로 조합 생성 + 무신사 상품수 조회.

        모든 브랜드(IP위험 포함)의 조합을 생성한다.
        existing_combos: 이미 존재하는 검색그룹 이름/키워드 set → 중복 제외.
        """
        all_brand_names = {bs.brand for bs in safe_brands}
        _existing = {s.lower() for s in (existing_combos or set())}

        # 모든 브랜드 쌍 + 점수순 정렬 (IP위험 포함)
        valid_pairs = [p for p in brand_keyword_pairs if p.brand in all_brand_names]
        valid_pairs.sort(key=lambda x: x.score, reverse=True)

        # 중복 제거 (같은 brand+keyword)
        seen_pairs: set[str] = set()
        unique_pairs: list[BrandKeywordPair] = []
        for p in valid_pairs:
            key = f"{p.brand}|{p.keyword}"
            if key not in seen_pairs:
                seen_pairs.add(key)
                unique_pairs.append(p)

        # pair가 없는 브랜드 → 카테고리 키워드와 자동 조합
        paired_brands = {p.brand for p in unique_pairs}
        all_cat_kws = set()
        for kws in CATEGORY_KEYWORDS.values():
            all_cat_kws.update(kws)

        for bs in safe_brands:
            if bs.brand in paired_brands:
                continue
            # 해당 브랜드의 카테고리에 맞는 키워드 선택
            cat_kws_for_brand: list[str] = []
            for cat in bs.categories or []:
                cat_kws_for_brand.extend(CATEGORY_KEYWORDS.get(cat, []))
            if not cat_kws_for_brand:
                # 카테고리 없으면 전체 키워드 중 주요 항목
                cat_kws_for_brand = [
                    "운동화",
                    "티셔츠",
                    "반팔",
                    "자켓",
                    "바람막이",
                    "슬리퍼",
                ]
            for kw in cat_kws_for_brand:
                key = f"{bs.brand}|{kw}"
                if key not in seen_pairs:
                    seen_pairs.add(key)
                    unique_pairs.append(
                        BrandKeywordPair(
                            brand=bs.brand,
                            keyword=kw,
                            source=bs.source,
                            count=1,
                            score=bs.score,
                        )
                    )

        combinations: list[Combination] = []

        # 무신사 API 병렬 조회 (세마포어로 동시 요청 제한)
        semaphore = asyncio.Semaphore(5)

        async def _fetch_count(pair: BrandKeywordPair):
            async with semaphore:
                try:
                    search_keyword = f"{pair.brand} {pair.keyword}"
                    result = await self.musinsa.search_products(
                        keyword=search_keyword,
                        sort="POPULAR",
                        size=1,
                    )
                    return pair, result.get("totalCount", 0), search_keyword
                except Exception as e:
                    logger.warning(
                        f"[AI소싱] 조합 조회 실패: {pair.brand}×{pair.keyword}: {e}"
                    )
                    return pair, 0, ""

        results = await asyncio.gather(*[_fetch_count(p) for p in unique_pairs])

        skipped_existing = 0
        for pair, total, search_keyword in results:
            if total == 0:
                continue

            # 기존 검색그룹 중복 체크
            if _existing:
                group_name = f"musinsa_{pair.brand}_{pair.keyword}".lower()
                kw_combo = f"{pair.brand} {pair.keyword}".lower()
                if group_name in _existing or kw_combo in _existing:
                    skipped_existing += 1
                    continue

            # 키워드가 속한 카테고리 찾기
            cat_name = ""
            cat_code = ""
            for cname, kws in CATEGORY_KEYWORDS.items():
                if pair.keyword in kws:
                    cat_name = cname
                    cat_code = MUSINSA_CATEGORIES.get(cname, "")
                    break

            search_url = (
                f"https://www.musinsa.com/search/goods"
                f"?keyword={search_keyword}&sort=pop"
            )

            combinations.append(
                Combination(
                    source_site="MUSINSA",
                    brand=pair.brand,
                    keyword=pair.keyword,
                    category=cat_name,
                    category_code=cat_code,
                    estimated_count=total,
                    search_url=search_url,
                )
            )

        # 예상상품수 내림차순
        combinations.sort(key=lambda c: c.estimated_count, reverse=True)

        total_est = sum(c.estimated_count for c in combinations)
        if skipped_existing > 0:
            logger.info(f"[AI소싱] 기존 검색그룹 중복 {skipped_existing}개 제외")
        logger.info(
            f"[AI소싱] 조합 생성 완료: {len(combinations)}개 그룹, "
            f"예상 {total_est:,}개 상품"
        )
        return combinations

    # ══════════════════════════════════════════
    # 6. 전체 분석 파이프라인
    # ══════════════════════════════════════════

    async def run_analysis(
        self,
        month: int = 0,
        main_category: str = "패션의류",
        excel_content: Optional[bytes] = None,
        target_count: int = 10000,
        forbidden_words: list[str] | None = None,
        use_musinsa: bool = True,
        use_naver: bool = True,
        musinsa_categories: list[str] | None = None,
        naver_categories: list[str] | None = None,
    ) -> dict[str, Any]:
        """전체 AI 소싱 분석 실행."""
        cat_map = MAIN_CATEGORIES.get(main_category, MAIN_CATEGORIES["패션의류"])
        m_cats = musinsa_categories or cat_map["musinsa"]
        n_cats = naver_categories or cat_map["naver"]

        all_brands: dict[str, BrandScore] = {}
        all_pairs: dict[str, BrandKeywordPair] = {}

        # 1) 무신사
        if use_musinsa:
            musinsa_brands, musinsa_pairs = await self.fetch_musinsa_popular(
                categories=m_cats
            )
            for bs in musinsa_brands:
                if bs.brand in all_brands:
                    existing = all_brands[bs.brand]
                    existing.count += bs.count
                    existing.score += bs.score
                    for cat in bs.categories:
                        if cat not in existing.categories:
                            existing.categories.append(cat)
                    for kw in bs.keywords:
                        if kw not in existing.keywords:
                            existing.keywords.append(kw)
                else:
                    all_brands[bs.brand] = bs
            for p in musinsa_pairs:
                key = f"{p.brand}|{p.keyword}"
                if key in all_pairs:
                    all_pairs[key].count += p.count
                    all_pairs[key].score += p.score
                else:
                    all_pairs[key] = p

        # 2) 네이버
        if use_naver:
            naver_brands, naver_pairs = await self.fetch_naver_datalab(
                naver_categories=n_cats, month=month
            )
            for bs in naver_brands:
                if bs.brand in all_brands:
                    existing = all_brands[bs.brand]
                    existing.count += bs.count
                    existing.score += bs.score * 0.5
                    for kw in bs.keywords:
                        if kw not in existing.keywords:
                            existing.keywords.append(kw)
                else:
                    all_brands[bs.brand] = bs
            for p in naver_pairs:
                key = f"{p.brand}|{p.keyword}"
                if key in all_pairs:
                    all_pairs[key].count += p.count
                    all_pairs[key].score += p.score * 0.5
                else:
                    all_pairs[key] = p

        # 3) 엑셀
        if excel_content:
            excel_brands = self.parse_sales_excel(excel_content)
            for bs in excel_brands:
                if bs.brand in all_brands:
                    existing = all_brands[bs.brand]
                    existing.count += bs.count
                    existing.score += bs.score * 2
                    existing.total_sales += bs.total_sales
                    existing.avg_profit_rate = bs.avg_profit_rate
                    for kw in bs.keywords:
                        if kw not in existing.keywords:
                            existing.keywords.append(kw)
                else:
                    bs.score *= 2
                    all_brands[bs.brand] = bs

        sorted_brands = sorted(all_brands.values(), key=lambda x: x.score, reverse=True)

        # 4) IP안전 필터
        brand_names = [bs.brand for bs in sorted_brands]
        safety = await self.check_brand_safety(brand_names, forbidden_words or [])
        safe_brands = [
            bs for bs in sorted_brands if safety.get(bs.brand, {}).get("is_safe", True)
        ]

        # 5) 조합 생성 — 실제 발견된 브랜드×키워드 쌍 기반
        pair_list = sorted(all_pairs.values(), key=lambda x: x.score, reverse=True)
        combinations = await self.generate_combinations(
            safe_brands[:50],
            pair_list,
            target_count=target_count,
        )

        for combo in combinations:
            brand_safety = safety.get(combo.brand, {})
            combo.is_safe = brand_safety.get("is_safe", True)
            combo.safety_reason = brand_safety.get("reason", "")

        total_estimated = sum(c.estimated_count for c in combinations)
        summary = {
            "total_brands_found": len(all_brands),
            "safe_brands": len(safe_brands),
            "unsafe_brands": len(sorted_brands) - len(safe_brands),
            "total_combinations": len(combinations),
            "total_estimated_products": total_estimated,
            "target_count": target_count,
        }

        return {
            "brands": [
                {
                    "brand": bs.brand,
                    "count": bs.count,
                    "score": round(bs.score, 1),
                    "total_sales": bs.total_sales,
                    "avg_profit_rate": round(bs.avg_profit_rate * 100, 1),
                    "categories": bs.categories,
                    "keywords": bs.keywords,
                    "source": bs.source,
                    "is_safe": safety.get(bs.brand, {}).get("is_safe", True),
                    "safety_reason": safety.get(bs.brand, {}).get("reason", ""),
                }
                for bs in sorted_brands[:100]
            ],
            "combinations": [
                {
                    "source_site": c.source_site,
                    "brand": c.brand,
                    "keyword": c.keyword,
                    "category": c.category,
                    "category_code": c.category_code,
                    "estimated_count": c.estimated_count,
                    "search_url": c.search_url,
                    "is_safe": c.is_safe,
                    "safety_reason": c.safety_reason,
                }
                for c in combinations
            ],
            "summary": summary,
        }
